import openpyxl
from openpyxl.utils import get_column_letter
from enum import Enum
import os
import logging
from typing import Optional

logging.basicConfig(level=logging.DEBUG)

class EChange(Enum):
    K_CNTT_KTD = "Khoa Công nghệ thông tin - Kỹ thuật điện"
    K_DL_KS = "Khoa Du lịch - Khách sạn"
    K_CK = "Khoa Cơ khí"
    K_KT_L = "Khoa Kinh tế- Luật"
    K_CSSD_NDT = "Khoa Chăm sóc sắc đẹp - Nuôi dưỡng trẻ"
    K_YD = "Khoa Y dược"
    K_NN = "Khoa Ngoại Ngữ"
    KHCB = "Khoa học cơ bản"

def copy_dates_and_add_columns(file_path: str) -> bool:
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        logging.debug(f"Processing {file_path}, max_row: {sheet.max_row}, max_column: {sheet.max_column}")

        # Kiểm tra ngày đã có sẵn từ cột 7
        for col in range(7, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            if isinstance(cell.value, str) and '/' in cell.value:
                logging.debug(f"Date {cell.value} already present at col {col}")

        # Bỏ logic sao chép ngày (vì đã có sẵn)
        # Bỏ logic thay "S", "C" thành "Buổi sáng", "Buổi chiều"
        # Bỏ việc chèn hàng 3 với "C1", "C2", ...

        # Log kết quả
        header_row = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        logging.debug(f"Header row after processing: {header_row}")
        row2 = [sheet.cell(row=2, column=col).value for col in range(1, sheet.max_column + 1)]
        logging.debug(f"Row 2 (sessions): {row2}")

        workbook.save(file_path)
        logging.info(f"Saved {file_path} after copy_dates_and_add_columns")
        return True
    except Exception as e:
        logging.error(f"Error in copy_dates_and_add_columns: {e}")
        return False

def summarize_k_attendance(file_path: str, faculty_name: str, output_file_path: Optional[str] = None) -> bool:
    if output_file_path is None:
        output_file_path = file_path
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        source_sheet = workbook.active
        logging.debug(f"Summarizing {file_path}, max_row: {source_sheet.max_row}, max_column: {source_sheet.max_column}")

        header_row = [source_sheet.cell(row=1, column=col).value for col in range(1, source_sheet.max_column + 1)]
        logging.debug(f"Input header row: {header_row}")
        row2 = [source_sheet.cell(row=2, column=col).value for col in range(1, source_sheet.max_column + 1)]
        logging.debug(f"Input row 2 (sessions): {row2}")

        # Tạo sheet "Thống kê nghỉ học"
        summary_sheet_name = "Thống kê nghỉ học"
        if summary_sheet_name in workbook.sheetnames:
            workbook.remove(workbook[summary_sheet_name])
        summary_sheet = workbook.create_sheet(summary_sheet_name)
        logging.debug(f"Created sheet: {summary_sheet_name}")

        headers = ["Ngày", "Họ và tên HSSV", "Khoa", "Lớp", "Giáo viên giảng dạy", "Nề nếp", "Buổi", "Phòng"]
        for col, header in enumerate(headers, 1):
            summary_sheet[f"{get_column_letter(col)}1"] = header
        logging.debug("Added headers to summary sheet")

        workbook.save(output_file_path)
        logging.debug(f"Saved {output_file_path} with empty summary sheet")

        # Thống kê dữ liệu
        class_name = os.path.basename(file_path).replace('.xlsx', '')
        current_row = 2

        for row in range(3, source_sheet.max_row + 1):  # Bắt đầu từ hàng 3 vì không có hàng "C1", "C2", ...
            ho_dem = source_sheet[f"C{row}"].value or ""
            ten = source_sheet[f"D{row}"].value or ""
            full_name = f"{ho_dem} {ten}".strip()
            
            for base_col in range(7, source_sheet.max_column + 1, 4):  # Duyệt từng nhóm 4 cột
                date_value = source_sheet.cell(row=1, column=base_col).value
                if isinstance(date_value, str) and '/' in date_value:
                    for offset in range(0, 4):  # Duyệt C1, C2, C3, C4 trong nhóm
                        col = base_col + offset
                        if col > source_sheet.max_column:
                            break
                        cell_value = source_sheet.cell(row=row, column=col).value
                        if cell_value == "K":
                            # Xác định buổi dựa trên vị trí cột
                            session = "Buổi sáng" if offset in [0, 1] else "Buổi chiều"  # C1, C2 là buổi sáng; C3, C4 là buổi chiều
                            logging.debug(f"Found 'K' at row {row}, col {col}: {full_name}, {date_value}, {session}")
                            summary_sheet[f"A{current_row}"] = date_value
                            summary_sheet[f"B{current_row}"] = full_name
                            summary_sheet[f"C{current_row}"] = faculty_name
                            summary_sheet[f"D{current_row}"] = class_name
                            summary_sheet[f"E{current_row}"] = ""
                            summary_sheet[f"F{current_row}"] = "Nghỉ học"
                            summary_sheet[f"G{current_row}"] = session
                            summary_sheet[f"H{current_row}"] = ""
                            current_row += 1

        for col in range(1, 9):
            summary_sheet.column_dimensions[get_column_letter(col)].width = 20

        workbook.save(output_file_path)
        logging.info(f"Saved {output_file_path} with summary sheet, rows added: {current_row - 2}")
        return True
    except Exception as e:
        logging.error(f"Error in summarize_k_attendance: {e}")
        return False