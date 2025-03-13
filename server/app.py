from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from handleExcel import copy_dates_and_add_columns, summarize_k_attendance
from uploadGgSheet import extract_spreadsheet_id, upload_to_google_sheets, connect_to_google_sheets
from network_checker import check_network, check_internet_speed
import tempfile
import openpyxl
from dotenv import load_dotenv

import logging
import shutil

load_dotenv()
CREDENTIALS_FILE = os.getenv('GOOGLE_CREDENTIALS_PATH', 'credentials.json')

logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)
CORS(app, resources={r"/check-network": {"origins": "*"}, r"/process": {"origins": "*"}})

@app.route('/', methods=['GET'])
def home():
    return jsonify({"message": "Welcome to Excel to Google Sheets Uploader API."})

@app.route('/check-network', methods=['POST', 'GET'])
def check_network_status():
    is_connected, network_message = check_network()
    if not is_connected:
        return jsonify({"error": True, "message": network_message})
    
    is_speed_ok, speed_message = check_internet_speed()
    if not is_speed_ok:
        return jsonify({"error": True, "message": speed_message})
    
    message = f"{network_message} - {speed_message}" if "Mạng chậm" in network_message or "Mạng chậm" in speed_message else "Mạng ổn định."
    logging.info(message)
    return jsonify({"error": False, "message": message})

@app.route('/process', methods=['POST'])
def process_files():
    files = request.files.getlist('files')
    google_sheet_url = request.form.get('googleSheetUrl')
    sheet_name = request.form.get('sheetName')
    faculty_name = request.form.get('faculty')

    if not files or not google_sheet_url or not sheet_name or not faculty_name:
        return jsonify({"status": "Lỗi: Thiếu thông tin!", "error": True})

    spreadsheet_id = extract_spreadsheet_id(google_sheet_url)
    if not spreadsheet_id:
        return jsonify({"status": "Lỗi: URL Google Sheet không hợp lệ!", "error": True})

    service = connect_to_google_sheets(spreadsheet_id)
    if not service:
        return jsonify({"status": "Lỗi: Không kết nối được Google Sheets!", "error": True})

    successful_files = 0
    total_rows_added = 0
    temp_dir = tempfile.mkdtemp()
    errors = []

    for file in files:
        if not file.filename.lower().endswith('.xlsx'):
            errors.append(f"Bỏ qua '{file.filename}': Không phải Excel.")
            continue

        file_path = os.path.join(temp_dir, file.filename)
        file.save(file_path)
        logging.debug(f"Saved file: {file_path}")

        if not copy_dates_and_add_columns(file_path):
            errors.append(f"Lỗi: Không xử lý được ngày và cột cho '{file.filename}'.")
            continue

        is_valid, error_message = validate_excel_data(file_path)
        if not is_valid:
            errors.append(f"Lỗi file '{file.filename}': {error_message}")
            continue

        rows_added = process_single_file(file_path, spreadsheet_id, sheet_name, faculty_name, service)
        if rows_added is not False:
            successful_files += 1
            total_rows_added += rows_added
        else:
            errors.append(f"Lỗi xử lý file '{file.filename}'.")

    shutil.rmtree(temp_dir)
    status = f"Xử lý hoàn tất: {successful_files} file thành công, {total_rows_added} hàng."
    response = {"status": status, "successfulFiles": successful_files, "totalRowsAdded": total_rows_added, "error": False}
    if errors:
        response["errors"] = errors
    logging.info(status)
    return jsonify(response)

def validate_excel_data(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        if sheet.max_row < 4:
            return False, "File cần ít nhất 4 hàng."
        for row in range(4, sheet.max_row + 1):
            if not sheet[f'C{row}'].value or not sheet[f'D{row}'].value:
                return False, f"Thiếu họ tên ở hàng {row}."
        
        has_date = False
        for col in range(7, sheet.max_column + 1):
            date_value = sheet.cell(row=1, column=col).value
            if isinstance(date_value, str) and '/' in date_value:
                has_date = True
                break
        if not has_date:
            return False, "Không tìm thấy ngày hợp lệ."
        
        return True, ""
    except Exception as e:
        return False, f"Lỗi kiểm tra dữ liệu: {str(e)}"

def process_single_file(file_path, spreadsheet_id, sheet_name, faculty_name, service):
    try:
        if not summarize_k_attendance(file_path, faculty_name):
            logging.error(f"Failed to summarize attendance for {file_path}")
            return False

        workbook = openpyxl.load_workbook(file_path)
        if "Thống kê nghỉ học" not in workbook.sheetnames:
            logging.error(f"Sheet 'Thống kê nghỉ học' not found in {file_path} after summarize")
            return False
        
        summary_sheet = workbook["Thống kê nghỉ học"]
        rows_added = summary_sheet.max_row - 1  # Trừ header
        logging.debug(f"Summary sheet has {rows_added} rows of data")
        
        if not upload_to_google_sheets(file_path, spreadsheet_id, sheet_name, service):
            logging.error(f"Failed to upload {file_path}")
            return False
        
        logging.debug(f"Processed {file_path}: {rows_added} rows.")
        return rows_added
    except Exception as e:
        logging.error(f"Error processing {file_path}: {e}")
        return False

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)