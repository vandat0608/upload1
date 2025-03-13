import os
import logging
import json
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from typing import Optional, List, Tuple
from dotenv import load_dotenv

# Load biến môi trường từ file .env
load_dotenv()

# Cấu hình logging
logging.basicConfig(level=logging.DEBUG)  # Đặt level DEBUG để ghi chi tiết

# Định nghĩa CREDENTIALS_FILE sử dụng biến môi trường hoặc đường dẫn tĩnh
CREDENTIALS_FILE = os.getenv('GOOGLE_CREDENTIALS_PATH', 'credentials.json')
logging.debug(f"Using credentials file: {CREDENTIALS_FILE}")

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

def extract_spreadsheet_id(url: str) -> Optional[str]:
    """
    Trích xuất ID của Google Spreadsheet từ URL.
    """
    if not url or not isinstance(url, str):
        logging.error("URL is invalid or empty")
        return None
    try:
        if 'spreadsheets/d/' in url:
            start = url.find('spreadsheets/d/') + len('spreadsheets/d/')
            end = url.find('/edit', start)
            if end == -1:
                end = url.find('/', start)
            if start != -1 and end != -1:
                spreadsheet_id = url[start:end].strip()
                logging.debug(f"Extracted spreadsheet ID: {spreadsheet_id}")
                return spreadsheet_id
        logging.error("URL format invalid, missing 'spreadsheets/d/'")
        return None
    except Exception as e:
        logging.error(f"Error extracting spreadsheet ID: {e}")
        return None

def connect_to_google_sheets(spreadsheet_id: str):
    """
    Thực hiện kết nối với google sheet
    """
    try:
        creds_json = os.getenv('GOOGLE_CREDENTIALS')
        if creds_json:
            logging.debug("Loading credentials from environment variable")
            credentials_data = json.loads(creds_json)
            creds = service_account.Credentials.from_service_account_info(
                credentials_data, scopes=SCOPES
            )
        elif os.path.exists(CREDENTIALS_FILE):
            logging.debug(f"Loading credentials from file: {CREDENTIALS_FILE}")
            creds = service_account.Credentials.from_service_account_file(
                CREDENTIALS_FILE, scopes=SCOPES
            )
        else:
            raise FileNotFoundError(f"No credentials found in env or file: {CREDENTIALS_FILE}")

        service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
        logging.debug("Connected to Google Sheets API")
        return service
    except Exception as e:
        logging.error(f"Error connecting to Google Sheets: {e}")
        return None

def get_first_empty_row(service, spreadsheet_id: str, sheet_name: str) -> int:
    """
    Tìm hàng trống tiếp theo trong Google Sheets, kiểm tra từ cột A đến H.
    """
    try:
        range_name = f'{sheet_name}!A:H'
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=range_name
        ).execute()
        
        values = result.get('values', [])
        if not values:
            logging.debug("No data in Google Sheets, starting from row 1")
            return 1  # Bắt đầu từ hàng 1 nếu chưa có dữ liệu
        
        # Tìm hàng trống đầu tiên (tất cả cột A-H đều trống)
        for row_idx, row in enumerate(values, 1):
            if not any(cell for cell in row[:8]):  # Kiểm tra 8 cột A-H
                logging.debug(f"Found empty row at {row_idx}")
                return row_idx
        # Nếu không có hàng trống, trả về hàng tiếp theo sau dữ liệu cuối
        next_row = len(values) + 1
        logging.debug(f"No empty row found, appending at row {next_row}")
        return next_row
    except HttpError as e:
        logging.error(f"Error checking data on Google Sheets: {e}")
        return 1  # Mặc định từ hàng 1 nếu lỗi

def upload_to_google_sheets(file_path, spreadsheet_id, sheet_name, service) -> bool:
    """
    Upload dữ liệu từ sheet 'Thống kê nghỉ học' lên Google Sheets, ánh xạ đúng cột A-H.
    """
    try:
        # Kiểm tra sheet "Thống kê nghỉ học"
        workbook = openpyxl.load_workbook(file_path)
        if "Thống kê nghỉ học" not in workbook.sheetnames:
            logging.error(f"Sheet 'Thống kê nghỉ học' not found in {file_path}")
            return False
        
        sheet = workbook["Thống kê nghỉ học"]
        data_to_upload = []

        # Đọc dữ liệu từ sheet "Thống kê nghỉ học"
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ hàng header
            date_value = str(row[0]) if row[0] is not None else ""  # Cột A: Ngày
            full_name = str(row[1]) if row[1] is not None else ""   # Cột B: Họ và tên HSSV
            faculty = str(row[2]) if row[2] is not None else ""     # Cột C: Khoa
            class_name = str(row[3]) if row[3] is not None else ""  # Cột D: Lớp
            teacher = str(row[4]) if row[4] is not None else ""     # Cột E: Giáo viên giảng dạy
            attendance = str(row[5]) if row[5] is not None else ""  # Cột F: Nề nếp
            session = str(row[6]) if row[6] is not None else ""     # Cột G: Buổi
            room = str(row[7]) if row[7] is not None else ""        # Cột H: Phòng

            # Gộp "Nề nếp" và "Buổi" thành một chuỗi cho cột H
            combined_attendance = f"{attendance} {session}".strip() if session else attendance

            # Ánh xạ dữ liệu theo thứ tự cột A-H của Google Sheets
            mapped_row = [
                "",              # Cột A: Để trống
                date_value,      # Cột B: Ngày
                room,            # Cột C: Phòng
                full_name,       # Cột D: Họ và tên HSSV
                faculty,         # Cột E: Khoa
                class_name,      # Cột F: Lớp
                teacher,         # Cột G: Giáo viên giảng dạy
                combined_attendance  # Cột H: Nề nếp Buổi
            ]
            data_to_upload.append(mapped_row)
        
        if not data_to_upload:
            logging.warning(f"No data to upload from {file_path}")
            return True  # Vẫn trả về True nếu không có dữ liệu để upload

        # Thêm header nếu cần
        headers = ["", "Ngày", "Phòng", "Họ và tên HSSV", "Khoa", "Lớp", "Giáo viên giảng dạy", "Nề nếp Buổi"]
        start_row = get_first_empty_row(service, spreadsheet_id, sheet_name)
        
        # Nếu sheet trống, thêm header vào hàng 1 và dữ liệu từ hàng 2
        if start_row == 1:
            data_to_upload.insert(0, headers)
            start_row = 1
        else:
            # Nếu đã có dữ liệu, chỉ thêm dữ liệu mới vào hàng trống tiếp theo
            start_row = max(start_row, 2)  # Đảm bảo không ghi đè header

        # Upload dữ liệu
        range_name = f"{sheet_name}!A{start_row}"
        body = {"values": data_to_upload}
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption="RAW",
            body=body
        ).execute()
        
        logging.info(f"Successfully uploaded {file_path} to Google Sheets starting at row {start_row}")
        return True
    except Exception as e:
        logging.error(f"Error uploading {file_path} to Google Sheets: {e}")
        return False

# Xóa các hàm không cần thiết (read_excel_data, push_data_to_google_sheets) để đơn giản hóa